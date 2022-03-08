# -*- coding: utf-8 -*-
"""
Created to ease processing Excel files.

@author: Carlos
"""

import os
import csv
import xlrd
import openpyxl
import warnings

warnings.simplefilter("ignore")


class ExcelFolder:
    # Class to represent a bunch of excel files in a folder. Those files
    # will be processed as desire using also the ExcelFile class


    # Method: Constructor
    # Specify the path to a folder where the Excel files are. 
    # By default is the path where the script is    
    def __init__(self, folderPath=os.path.abspath(os.path.dirname(__file__))):     
        
        if type(folderPath) == str and os.path.exists(folderPath): # Folder must be str and exists
        
            self.folderPath = folderPath
            self.excelFiles, self.badFiles = self.GetExcelFilesFromFolder()

        else: 
            raise Exception("Path value must be valid")


    # Method: GetFilesFromFolder
    # Get all the XLS files in the path
    def GetExcelFilesFromFolder(self): 
        
        folder = os.listdir(self.folderPath)
        pathExcelFiles=[]
        pathXlsBadFiles=[]
        valid_extensions = ['xls', 'xlsx', 'csv']
        
        for file in folder:        
            try: 
                extension = file.split('.')[1]
                
                if extension in valid_extensions:
                    try:
                        pathExcelFiles.append(self.GetExcelFile(f"{self.folderPath}/{file}"))
                    
                    except Exception:
                        pathXlsBadFiles.append(f"{self.folderPath}/{file}")
                        
            except IndexError: # To handle if there are files with no extension
                pass

                
                            
                    
        return list(pathExcelFiles), list(pathXlsBadFiles)


    # Method: GetExcelFile
    # Try to open the files with the xls and xlsx method
    def GetExcelFile(self, filePath): 
        try: 
            return XlsFile(filePath)
        
        except Exception: 
            
            try: 
                return XlsxFile(filePath)
            
            except Exception: 
                raise Exception (f"[X] ERROR - File {filePath} not supported") 
                            

    # Method: PrintFiles
    # Print all the files which we are working with
    def PrintGoodFiles(self): 
        for file in self.excelFiles:
            print(file)
    
    
    # Method: PrintBadFiles
    # Print all the files with some problem so they can't be processed
    def PrintBadFiles(self): 
        for file in self.badFiles:
            print(f"\n[X] Bad file: {file}")



class XlsFile:
    # Class to represent a single XLS file with its data


    # Method: Constructor
    # Specify the XLS file path to work with and the sheet number (starting in 0) 
    # By default the excel file will work with the first sheet.
    def __init__(self, filePath, sheetNumber=0):
        
        if os.path.isfile(filePath): 
            self.fileName = filePath
            self.excelWorkbook = xlrd.open_workbook(self.fileName)
            self.workingSheet = self.excelWorkbook.sheet_by_index(sheetNumber)
            self.rowsNumber= self.GetNumberOfRows()
            self.columnsNumber = self.GetNumberOfColumns()
                
        else: 
            raise Exception("File must exists and has to be provided to the constructor")
    
    # Method: Printer
    # Specify the way that an object of this class is printed whith print clausule
    def __str__(self):
        return f"\nFilename - {self.fileName}\n\tColumns: {self.columnsNumber}\n\tRows: {self.rowsNumber}"


    # Method: Instancer?
    # Change how to display info when using the type() clausule
    def __repr__(self):
        return f"XlsFile - {self.fileName}"
    
    
    # Method: GetNumberOfRows
    # Get number of rows from the Excel sheet
    def GetNumberOfRows(self): 
        return self.workingSheet.nrows
    

    # Method: GetNumberOfColumns
    # Get number of columns from the Excel sheet
    def GetNumberOfColumns(self): 
        return self.workingSheet.ncols
    

    # Method: GetColumnData
    # Get all the data from a colum starting in the row startRow until the row "endRow"
    # (by defaultall the values of the column are returned)
    def GetColumnData(self, columNumber, startRow=0, endRow=None):
        
        if endRow == None: 
            endRow = self.rowsNumber
            
        return self.workingSheet.col_values(columNumber, start_rowx=startRow, end_rowx=endRow)
        

    # Method: GetRowData
    # Get all the data from a row starting in the row startRow until the row "endRow"
    # (by defaultall the values of the column are returned)
    def GetRowData(self, rowNumber, startCol=0, endCol=None):
        
        if endCol == None: 
            endCol = self.columnsNumber
            
        return self.workingSheet.row_values(rowNumber, start_colx=startCol, end_colx=endCol)
    

    # Method: GetAllDataByColumns
    # Get all the data from the columns specified from startCol to stopColumn. 
    # This method return an array of arrays. In each position of the array it's a whole column
    def GetAllDataByColumns(self, startColumn=0, stopColumn=None):
        
        columnsData = []
        
        if stopColumn == None: 
            stopColumn = self.columnsNumber
            
        for column in range(startColumn, stopColumn):
            columnsData.append(self.GetColumnData(column))
            
        return columnsData
            
            
    # Method: GetAllDataByRows
    # Get all the data from the rows specified from startRow to stopRow. 
    # This method return an array of arrays. In each position of the array it's a whole row
    def GetAllDataByRows(self, startRow=0, stopRow=None):
        
        rowsData = []
        
        if stopRow == None: 
            stopRow = self.rowsNumber
        
        for row in range(startRow, stopRow):
            rowsData.append(self.GetRowData(row))
            
        return rowsData
    
    
    # Method: changeWorkSheet
    # Change the current worksheet to work with
    def ChangeWorkingSheet(self, sheetNumber):
        try: 
            self.workingSheet = (self.excelWorkbook).sheet_by_index(sheetNumber)
            
        except Exception: 
            raise Exception("ERROR when trying to change the working sheet. " +
                            "To change the working sheet you must provide a valid index (starting in 0)")



class XlsxFile:
    # Class to represent a single XLSX file with its data


    # Method: Constructor
    # Specify the XLSX file path to work with and the sheet number (starting in 0) 
    # By default the excel file will work with the first sheet.
    def __init__(self, filePath, sheetNumber=0):
        
        if os.path.isfile(filePath): 
            self.fileName = filePath
            self.excelWorkbook = openpyxl.load_workbook(filename=self.fileName, data_only=True)#, read_only=True)
            self.workingSheet = self.excelWorkbook.worksheets[sheetNumber]
            self.rowsNumber= self.GetNumberOfRows()
            self.columnsNumber = self.GetNumberOfColumns()
            
        else: 
            raise Exception("File must exists and has to be provided to the constructor")
    
    # Method: Printer
    # Specify the way that an object of this class is printed whith print clausule
    def __str__(self):
        return f"\nFilename - {self.fileName}\n\tColumns: {self.columnsNumber}\n\tRows: {self.rowsNumber}"


    # Method: Instancer?
    # Change how to display info when using the type() clausule
    def __repr__(self):
        return f"XlsxFile - {self.fileName}"
    
    
    # Method: GetNumberOfRows
    # Get number of rows from the Excel sheet
    def GetNumberOfRows(self): 
        return self.workingSheet.max_row
    

    # Method: GetNumberOfColumns
    # Get number of columns from the Excel sheet
    def GetNumberOfColumns(self): 
        return self.workingSheet.max_column
    

    # Method: GetColumnData
    # Get all the data from a colum starting in the row startRow until the row "endRow"
    # (by defaultall the values of the column are returned)
    def GetColumnData(self, columNumber, startRow=0, endRow=None):
        
        columNumber += 1 #Adjusting:  xlrd begins in 0 while openpyxl begins in 1
        if endRow == None: 
            endRow = self.rowsNumber
            
        rowData = list(self.workingSheet.iter_cols(min_col=columNumber, max_col=columNumber, min_row=startRow, max_row=endRow, values_only=True))[0]
        return list(rowData)

        

    # Method: GetRowData
    # Get all the data from a row starting in the column startCol until the column "endCol"
    # (by defaultall the values of the column are returned)
    def GetRowData(self, rowNumber, startCol=0, endCol=None):
        
        rowNumber += 1 #Adjusting:  xlrd begins in 0 while openpyxl begins in 1
        if endCol == None: 
            endCol = self.columnsNumber
            
        rowData = list(self.workingSheet.iter_rows(min_row=rowNumber, max_row=rowNumber, min_col=startCol, max_col=endCol, values_only=True))[0]
        return list(rowData)


    # Method: GetAllDataByColumns
    # Get all the data from the columns specified from startCol to stopColumn. 
    # This method return an array of arrays. In each position of the array it's a whole column
    def GetAllDataByColumns(self, startColumn=0, stopColumn=None):
        
        columnsData = []
        
        if stopColumn == None: 
            stopColumn = self.GetNumberOfColumns()
            
        for column in range(startColumn, stopColumn):
            columnsData.append(self.GetColumnData(column))
            
        return columnsData
            
            
    # Method: GetAllDataByRows
    # Get all the data from the rows specified from startRow to stopRow. 
    # This method return an array of arrays. In each position of the array it's a whole row
    def GetAllDataByRows(self, startRow=0, stopRow=None):
        
        rowsData = []
        
        if stopRow == None: 
            stopRow = self.GetNumberOfRows()
        
        for row in range(startRow, stopRow):
            rowsData.append(self.GetRowData(row))
            
        return rowsData
    
    
    # Method: changeWorkSheet
    # Change the current worksheet to work with
    def ChangeWorkingSheet(self, sheetNumber):
        try: 
            self.workingSheet = self.excelWorkbook.worksheets[sheetNumber]
            
        except Exception: 
            raise Exception("ERROR when trying to change the working sheet. " +
                            "To change the working sheet you must provide a valid index (starting in 0)")




class CsvFile:
    # Class to represent a single XLSX file with its data


    # Method: Constructor
    # Specify the XLSX file path to work with and the sheet number (starting in 0) 
    # By default the excel file will work with the first sheet.
    def __init__(self, filePath, userDelimiter=','):
        
        if os.path.isfile(filePath): 
            self.fileName = filePath
            self.excelWorkbook = None
            self.workingSheet = None
            self.columnDelimiter = userDelimiter
            self.columnsNumber, self.rowsNumber, self.rowsData = self.GetNumberOfColumnsAndRows()            
            
        else: 
            raise Exception("File must exists and has to be provided to the constructor")


    # Method: Printer
    # Specify the way that an object of this class is printed whith print clausule
    def __str__(self):
        return f"\nFilename - {self.fileName}\n\tColumns: {self.columnsNumber}\n\tRows: {self.rowsNumber}"


    # Method: Instancer?
    # Change how to display info when using the type() clausule
    def __repr__(self):
        return f"CsvFile - {self.fileName}"
    
    
    # Method: GetNumberOfRows
    # Get number of rows from the Excel sheet
    def GetNumberOfColumnsAndRows(self): 
        
        with open(self.fileName, 'r') as file:
            self.workingSheet = csv.reader(file, delimiter=self.columnDelimiter)
    
            numberColumns = 0
            numberRows = 0
            rowsData =[]
            
            for row in self.workingSheet: 
                if len(row) > numberColumns: numberColumns = len(row)
                numberRows += 1
                rowsData.append(row)
                
        return numberColumns, numberRows, rowsData

 
    # Method: GetColumnData
    # Get all the data from a colum starting in the row startRow until the row "endRow"
    # (by defaultall the values of the column are returned)
    def GetColumnData(self, columnNumber, startRow=0, endRow=None):
        
        if endRow == None: 
            endRow = self.rowsNumber
    
        columnData = []            
        for row in self.rowsData:
            columnData.append(row[columnNumber])
            
        return columnData

        

    # Method: GetRowData
    # Get all the data from a row starting in the column startCol until the column "endCol"
    # (by defaultall the values of the column are returned)
    def GetRowData(self, rowNumber, startCol=0, endCol=None):
        
        if endCol == None: 
            endCol = self.columnsNumber
            
        return self.rowsData[rowNumber][startCol:endCol]


    # Method: GetAllDataByColumns
    # Get all the data from the columns specified from startCol to stopColumn. 
    # This method return an array of arrays. In each position of the array it's a whole column
    def GetAllDataByColumns(self, startColumn=0, stopColumn=None):
        
        columnsData = []
        
        if stopColumn == None: 
            stopColumn = self.GetNumberOfColumns()
            
        for column in range(startColumn, stopColumn):
            columnsData.append(self.GetColumnData(column))
            
        return columnsData
            
            
    # Method: GetAllDataByRows
    # Get all the data from the rows specified from startRow to stopRow. 
    # This method return an array of arrays. In each position of the array it's a whole row
    def GetAllDataByRows(self, startRow=0, stopRow=None):
        
        rowsData = []
        
        if stopRow == None: 
            stopRow = self.GetNumberOfRows()
        
        for row in range(startRow, stopRow):
            rowsData.append(self.GetRowData(row))
            
        return rowsData
    
    
    # Method: changeWorkSheet
    # Change the current worksheet to work with
    def ChangeWorkingSheet(self, sheetNumber):
        raise Exception("ERROR - CSV files has no sheets'")


